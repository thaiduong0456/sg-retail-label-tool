from __future__ import annotations

import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO, Iterable

import fitz
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


MASTER_ALIASES = {
    "sku": {"sku", "itemcode", "item code", "item#", "item no", "product code"},
    "description": {"itemdescription", "item description", "description", "product name", "product name in english"},
    "currency": {"currency", "currency symbol"},
    "rrp": {"rrp", "rrp (sgd)", "price", "retail price", "amount", "số tiền", "gia", "giá"},
    "ean": {"ean", "barcode", "upc", "barcode/upc"},
}

REQUEST_ALIASES = {
    "ean": MASTER_ALIASES["ean"],
    "co": {"co", "coo", "country", "country of origin", "origin", "xuất xứ", "xuat xu"},
    "sku": MASTER_ALIASES["sku"],
}

COUNTRY_NAMES = {
    "CN": "China",
    "CHN": "China",
    "CHINA": "China",
    "VN": "Vietnam",
    "VNM": "Vietnam",
    "VIETNAM": "Vietnam",
    "KH": "Cambodia",
    "KHM": "Cambodia",
    "CAMBODIA": "Cambodia",
    "IN": "India",
    "IND": "India",
    "INDIA": "India",
}


class InputError(ValueError):
    pass


@dataclass(frozen=True)
class LabelRecord:
    sku: str
    description: str
    currency: str
    rrp: str
    ean: str
    country: str
    source_row: int


@dataclass(frozen=True)
class MatchResult:
    labels: list[LabelRecord]
    missing: list[dict[str, str | int]]
    duplicate_master_eans: list[str]


def _key(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def normalize_identifier(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".15g")
    text = str(value).strip().replace("\u200e", "").replace("\u200f", "")
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    if re.fullmatch(r"\d+(?:\.\d+)?[Ee][+-]?\d+", text):
        try:
            number = float(text)
            if number.is_integer():
                return str(int(number))
        except ValueError:
            pass
    return text


def _find_header(ws, aliases: dict[str, set[str]], required: set[str], max_rows: int = 30):
    sheet_max_row, sheet_max_col = _sheet_bounds(ws)
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=min(sheet_max_row, max_rows),
            min_col=1,
            max_col=sheet_max_col,
            values_only=True,
        ),
        start=1,
    ):
        values = list(row)
        mapping: dict[str, int] = {}
        for col_idx, value in enumerate(values, start=1):
            normalized = _key(value)
            for field, names in aliases.items():
                if normalized in names and field not in mapping:
                    mapping[field] = col_idx
        if required.issubset(mapping):
            return row_idx, mapping
    wanted = ", ".join(sorted(required))
    raise InputError(f"Không tìm thấy dòng tiêu đề có đủ cột bắt buộc: {wanted}.")


def _sheet_bounds(ws) -> tuple[int, int]:
    """Return reliable bounds even when an XLSX omits dimension metadata."""
    cells = getattr(ws, "_cells", {})
    if cells:
        return max(cell.row for cell in cells.values()), max(cell.column for cell in cells.values())
    max_row = ws.max_row if isinstance(ws.max_row, int) else 0
    max_col = ws.max_column if isinstance(ws.max_column, int) else 0
    return max_row, max_col


def _open_workbook(source: str | Path | bytes | BinaryIO):
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    # Normal mode is more compatible with XLSX writers that omit worksheet
    # dimension metadata; the expected Master files are small enough for this.
    return load_workbook(source, read_only=False, data_only=True)


def _select_sheet(wb, aliases, required):
    errors = []
    for ws in wb.worksheets:
        try:
            header_row, columns = _find_header(ws, aliases, required)
            return ws, header_row, columns
        except InputError as exc:
            errors.append(f"{ws.title}: {exc}")
    raise InputError("Không có sheet phù hợp. " + " | ".join(errors))


def read_master(source: str | Path | bytes | BinaryIO) -> tuple[dict[str, dict], list[str]]:
    wb = _open_workbook(source)
    ws, header_row, columns = _select_sheet(wb, MASTER_ALIASES, {"ean", "sku", "description", "rrp"})
    records: dict[str, dict] = {}
    duplicates: list[str] = []
    sheet_max_row, sheet_max_col = _sheet_bounds(ws)
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            max_row=sheet_max_row,
            min_col=1,
            max_col=sheet_max_col,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        def value(field: str, default=None):
            column = columns.get(field)
            return row[column - 1] if column and column <= len(row) else default

        ean = normalize_identifier(value("ean"))
        if not ean:
            continue
        record = {
            "sku": normalize_identifier(value("sku")),
            "description": normalize_identifier(value("description")),
            "currency": normalize_identifier(value("currency")) if columns.get("currency") else "SGD",
            "rrp": value("rrp"),
            "ean": ean,
        }
        if ean in records:
            duplicates.append(ean)
            continue
        records[ean] = record
    if not records:
        raise InputError("Master data không có dòng dữ liệu EAN hợp lệ.")
    return records, sorted(set(duplicates))


def read_request(source: str | Path | bytes | BinaryIO) -> list[dict]:
    wb = _open_workbook(source)
    ws, header_row, columns = _select_sheet(wb, REQUEST_ALIASES, {"ean"})
    rows = []
    sheet_max_row, sheet_max_col = _sheet_bounds(ws)
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=header_row + 1,
            max_row=sheet_max_row,
            min_col=1,
            max_col=sheet_max_col,
            values_only=True,
        ),
        start=header_row + 1,
    ):
        def value(field: str, default=None):
            column = columns.get(field)
            return row[column - 1] if column and column <= len(row) else default

        ean = normalize_identifier(value("ean"))
        if not ean:
            continue
        rows.append({
            "ean": ean,
            "co": normalize_identifier(value("co")) if columns.get("co") else "",
            "sku": normalize_identifier(value("sku")) if columns.get("sku") else "",
            "row": row_idx,
        })
    if not rows:
        raise InputError("File EAN không có EAN hợp lệ.")
    return rows


def _format_price(value: object) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return f"{value:,.0f}" if float(value).is_integer() else f"{value:,.2f}".rstrip("0").rstrip(".")
    return str(value).strip()


def match_labels(master_source, request_source) -> MatchResult:
    master, duplicates = read_master(master_source)
    requests = read_request(request_source)
    labels: list[LabelRecord] = []
    missing: list[dict[str, str | int]] = []
    for requested in requests:
        item = master.get(requested["ean"])
        if not item:
            missing.append(requested)
            continue
        country = requested["co"].strip().upper()
        labels.append(LabelRecord(
            sku=item["sku"],
            description=item["description"],
            currency=item["currency"] or "SGD",
            rrp=_format_price(item["rrp"]),
            ean=item["ean"],
            country=country,
            source_row=requested["row"],
        ))
    return MatchResult(labels, missing, duplicates)


def _font_candidates():
    project_root = Path(__file__).resolve().parent
    return [
        ("Calibri", project_root / "assets/Calibri.ttf", project_root / "assets/Calibrib.ttf"),
        ("Calibri", Path("/Applications/Microsoft Word.app/Contents/Resources/DFonts/Calibri.ttf"), Path("/Applications/Microsoft Word.app/Contents/Resources/DFonts/Calibrib.ttf")),
        ("Calibri", Path("/Applications/Microsoft Excel.app/Contents/Resources/DFonts/Calibri.ttf"), Path("/Applications/Microsoft Excel.app/Contents/Resources/DFonts/Calibrib.ttf")),
        ("Calibri", Path("/Library/Fonts/Calibri.ttf"), Path("/Library/Fonts/Calibri Bold.ttf")),
        ("Carlito", project_root / "assets/Carlito-Regular.ttf", project_root / "assets/Carlito-Bold.ttf"),
        ("Arial", Path("/Library/Fonts/Arial.ttf"), Path("/Library/Fonts/Arial Bold.ttf")),
        ("Arial", Path("/System/Library/Fonts/Supplemental/Arial.ttf"), Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf")),
        ("Carlito", Path("/usr/share/fonts/truetype/crosextra/Carlito-Regular.ttf"), Path("/usr/share/fonts/truetype/crosextra/Carlito-Bold.ttf")),
    ]


def get_font_status() -> tuple[str, bool]:
    for name, regular, bold in _font_candidates():
        if regular.exists() and bold.exists():
            return name, name == "Calibri"
    return "Helvetica", False


def _register_font() -> tuple[str, str]:
    for name, regular, bold in _font_candidates():
        if regular.exists() and bold.exists():
            regular_name, bold_name = f"{name}-Label", f"{name}-Label-Bold"
            if regular_name not in pdfmetrics.getRegisteredFontNames():
                pdfmetrics.registerFont(TTFont(regular_name, str(regular)))
                pdfmetrics.registerFont(TTFont(bold_name, str(bold)))
            return regular_name, bold_name
    return "Helvetica", "Helvetica-Bold"


def _wrap_text(text: str, font: str, size: float, max_width: float) -> list[str]:
    words = str(text).split()
    if not words:
        return [""]
    lines, current = [], words[0]
    for word in words[1:]:
        trial = f"{current} {word}"
        if pdfmetrics.stringWidth(trial, font, size) <= max_width:
            current = trial
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def _draw_sg_label(c, record: LabelRecord, regular: str) -> None:
    """Draw one logical 4 x 3 cm SG label at the current origin."""
    width, height = 4.0 * cm, 3.0 * cm
    margin = 2.0
    x0, y0 = margin, margin
    table_w, table_h = width - 2 * margin, height - 2 * margin
    label_w = 21.0
    value_w = table_w - label_w
    # Proportions follow the supplied SG reference label.
    row_heights = [14.0, 14.0, table_h - 14.0 - 14.0 - 14.0 - 10.5, 14.0, 10.5]
    labels = ["EAN", "Item\nCode", "Item\nDesc.", "RRP\n(SGD)", "CO"]
    values = [record.ean, record.sku, record.description, f"SGD {record.rrp}".strip(), record.country]

    c.setStrokeColor(colors.black)
    c.setLineWidth(0.35)
    c.rect(x0, y0, table_w, table_h)
    c.line(x0 + label_w, y0, x0 + label_w, y0 + table_h)
    y_top = y0 + table_h
    for idx, (label, value, row_h) in enumerate(zip(labels, values, row_heights)):
        y_bottom = y_top - row_h
        if idx:
            c.line(x0, y_top, x0 + table_w, y_top)

        label_lines = label.split("\n")
        leading = 5.4
        label_y = y_bottom + (row_h + len(label_lines) * leading) / 2 - leading + 0.4
        c.setFont(regular, 5.0)
        for line in label_lines:
            c.drawString(x0 + 1.5, label_y, line)
            label_y -= leading

        if idx == 2:
            lines = _wrap_text(str(value), regular, 5.0, value_w - 3.0)
            if len(lines) > 5:
                lines = lines[:5]
                lines[-1] = lines[-1].rstrip() + "..."
        else:
            lines = [str(value)]
        text_y = y_bottom + (row_h + len(lines) * leading) / 2 - leading + 0.4
        c.setFont(regular, 5.0)
        for line in lines:
            c.drawString(x0 + label_w + 1.5, text_y, line)
            text_y -= leading
        y_top = y_bottom


def create_label_pdf(records: Iterable[LabelRecord]) -> bytes:
    records = list(records)
    if not records:
        raise InputError("Không có tem hợp lệ để tạo PDF.")
    # Three identical 4 x 3 cm labels per 9 x 4 cm page, matching the SG sample.
    width, height = 9.0 * cm, 4.0 * cm
    output = io.BytesIO()
    c = canvas.Canvas(output, pagesize=(width, height), pageCompression=1)
    regular, _bold = _register_font()

    for record in records:
        for copy_index in range(3):
            c.saveState()
            c.translate((copy_index + 1) * 3.0 * cm, 0)
            c.rotate(90)
            _draw_sg_label(c, record, regular)
            c.restoreState()
        c.showPage()
    c.save()
    return output.getvalue()


def render_pdf_pages(pdf_bytes: bytes, scale: float = 3.0, max_pages: int | None = None) -> list[bytes]:
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = []
    count = len(doc) if max_pages is None else min(len(doc), max_pages)
    for page_idx in range(count):
        pix = doc[page_idx].get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        pages.append(pix.tobytes("png"))
    return pages
